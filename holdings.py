"""
Investment Manager Holdings Fetcher — SEC 13F + N-PORT (edgartools edition)
===========================================================================
Uses the free `edgartools` library (no API key, no rate limits).

Outputs:
  - One Excel per manager: <name>_top20_<asofdate>.xlsx
  - One combined Excel:    all_managers_top20_<rundate>.xlsx

Install:  pip install edgartools openpyxl matplotlib
"""

import csv
import io
import json as _json
import os
import threading
import time
import urllib.request
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime

from edgar import set_identity, Company
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

set_identity("Investment Manager Holdings Fetcher holdings@example.com")

MAX_DATE = "2025-12-31"
TOP_N    = 20

def _max_filing_date():
    """Compute filing date cutoff: MAX_DATE + 75 days.
    13F/NPORT filings are due ~45 days after quarter end.
    If user sets MAX_DATE=2025-12-31 (meaning Q4 2025),
    we search for filings filed through ~mid-March 2026."""
    from datetime import datetime, timedelta
    try:
        dt = datetime.strptime(MAX_DATE, "%Y-%m-%d")
        return (dt + timedelta(days=75)).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return MAX_DATE

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

MANAGERS_13F = {
    "Baupost Group":       "1061768",
    "Elliott Management":  "1791786",
    "Durable Capital":     "1798849",
    "Tiger Global":        "1167483",
    "Aspex Management":    "1768375",
    "D1 Capital":          "1747057",
    "Viking Global":       "1103804",
    "Arrowstreet Capital": "1164508",
    "CC&L Q ACWI":         "1596800",
    "Spyglass Capital":    "1654344",
    "Wellington Mgmt":     "902219",
}

MANAGERS_NPORT = {}

# ── Ticker aliases — maps old/wrong tickers to current ones ──────────────────
_TICKER_ALIASES = {
    "FB": "META", "ANTM": "ELV", "TWTR": "X", "DISCA": "WBD",
    "DISCK": "WBD", "VIAC": "PARA", "VIACA": "PARA", "KSU": "CP",
    "CLHB": "CLH",
}

def _apply_ticker_aliases(records):
    """Fix outdated tickers (e.g. FB→META) in a list of holding dicts."""
    for rec in records:
        t = (rec.get("ticker") or "").upper()
        if t in _TICKER_ALIASES:
            rec["ticker"] = _TICKER_ALIASES[t]

# ── CUSIP → Ticker fallback (for SPDR ETFs that omit tickers) ───────────────

CUSIP_TO_TICKER = {
    # ── Mega-cap tech ────────────────────────────────────────────────────────
    "594918104": "MSFT",  "037833100": "AAPL",  "67066G104": "NVDA",
    "02079K305": "GOOGL", "02079K107": "GOOG",  "023135106": "AMZN",
    "30303M102": "META",  "11135F101": "AVGO",  "882508104": "TSLA",
    "594918AD2": "MSFT",  "00507V109": "ADBE",  "46120E602": "INTU",
    "79466L302": "CRM",   "568259108": "MRVL",  "007903107": "AMD",
    "458140100": "INTC",  "457030104": "IBM",   "68389X105": "ORCL",
    "17275R102": "CSCO",  "87612E106": "TMUS",  "20030N101": "CMCSA",
    "92826C839": "V",     "571748102": "MA",    "585055106": "MELI",
    "00724F101": "ADSK",  "74762E102": "QCOM",  "872540109": "TXN",
    "98138H101": "WM",    "04271T100": "ANET",  "29786A106": "EQIX",
    # ── Financials ───────────────────────────────────────────────────────────
    "46625H100": "JPM",   "46625H218": "JPM",   "060505104": "BAC",
    "172967424": "C",     "949746101": "WFC",   "38141G104": "GS",
    "58933Y105": "MS",    "09247X101": "BLK",   "78462F103": "SPGI",
    "55354G100": "MSCI",  "629491105": "NDAQ",  "075887109": "BDX",
    "808513105": "SCHW",  "14913Q104": "CB",    "29364G103": "ENB",
    "053015103": "AXP",   "459200101": "IBM",   "256135203": "DFS",
    "635405101": "NEE",   "29250N105": "ENB",   "742718109": "PG",
    # ── Healthcare ───────────────────────────────────────────────────────────
    "478160104": "JNJ",   "91324P102": "UNH",
    "69608A108": "PFE",   "002824100": "ABBV",  "532457108": "LLY",
    "88160R101": "TMO",   "718172109": "PM",
    "57636Q104": "MRK",   "053332102": "AZN",   "000360206": "ABT",
    "09062X103": "BIIB",  "92532F100": "VRTX",  "125523100": "CI",
    "004489403": "ACE",   "00846U101": "GILD",  "464287457": "ISRG",
    "94974B100": "WELL",  "75886F107": "REGN",  "126650100": "CVS",
    # ── Consumer / Retail ────────────────────────────────────────────────────
    "931142103": "WMT",   "22160K105": "COST",  "693475105": "PEP",
    "172062101": "KO",    "552953101": "MO",    "02209S103": "MO",
    "761152107": "RES",   "500754106": "KR",
    "866743105": "SBUX",  "580135101": "MCD",   "902973304": "TGT",
    "037833108": "AAPL",  "913017109": "UPS",   "35137L105": "FDX",
    # ── Energy ───────────────────────────────────────────────────────────────
    "30231G102": "XOM",   "166764100": "CVX",   "192446102": "COP",
    "81369Y506": "SLB",   "302445101": "EOG",   "718507105": "PSX",
    "59001A102": "MPC",   "91911K102": "VLO",   "670346105": "OXY",
    "35671D857": "FANG",  "845467105": "SPH",
    # ── Industrials ──────────────────────────────────────────────────────────
    "443185106": "HON",   "149123101": "CAT",   "369604301": "GE",
    "20825C104": "COP",   "091166105": "BA",    "032511107": "AMGN",
    "912456100": "RTX",   "26441C204": "DUK",
    "025537101": "AEP",   "842587107": "SO",
    "524651106": "LECO",  "370334104": "GD",    "539830109": "LMT",
    # ── Telecom / Utilities / Insurance ──────────────────────────────────────
    "00206R102": "T",     "92343V104": "VZ",    "843034101": "NEE",
    "026874784": "AIG",   "743315103": "PGR",   "59156R108": "MET",
    "902494103": "TFC",
    "607059104": "MMM",   "88579Y101": "MMC",   "29089Q105": "EMR",
    "404280109": "HSIC",  "11133T103": "BMY",
    # ── Other large-cap ──────────────────────────────────────────────────────
    "025816109": "AMGN",  "624756102": "MUFG",
    "98978V103": "ZTS",   "904764200": "UBER",  "90384S303": "ULTA",
    "89832Q109": "TSCO",  "828806109": "SHW",   "48203R104": "JNPR",
    "68902V107": "OTIS",  "855244109": "SBAC",  "31428X106": "FE",
    "184496107": "CLH",   # Clean Harbors
    "025537101": "AEP",   # American Electric Power
}

# ── OpenFIGI CUSIP → Ticker resolution ───────────────────────────────────────

_figi_cache = {}
_figi_cache_lock = threading.Lock()


def openfigi_lookup(cusips):
    """
    Batch lookup CUSIPs via OpenFIGI API (free, no key needed).
    Anonymous API: 20 req/min, 100 items per request.
    Returns dict mapping CUSIP -> ticker (or 'N/A' if not found).
    """
    results = {}

    # Check cache first
    uncached = []
    with _figi_cache_lock:
        for cusip in cusips:
            if cusip in _figi_cache:
                results[cusip] = _figi_cache[cusip]
            else:
                uncached.append(cusip)

    if not uncached:
        return results

    # Batch into groups of 50 (API limit is 100, but smaller batches avoid 413 errors)
    for i in range(0, len(uncached), 50):
        batch = uncached[i:i + 50]
        payload = [{"idType": "ID_CUSIP", "idValue": c} for c in batch]
        for attempt in range(2):  # try twice
            try:
                req = urllib.request.Request(
                    "https://api.openfigi.com/v3/mapping",
                    data=_json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urllib.request.urlopen(req, timeout=20) as resp:
                    data = _json.loads(resp.read().decode("utf-8"))

                for j, item in enumerate(data):
                    cusip = batch[j]
                    ticker = "N/A"
                    if "data" in item and item["data"]:
                        for d in item["data"]:
                            t = d.get("ticker", "")
                            if t and t not in ("", "N/A"):
                                ticker = t
                                break
                    results[cusip] = ticker
                    with _figi_cache_lock:
                        _figi_cache[cusip] = ticker
                break  # success — exit retry loop
            except Exception as e:
                if attempt == 0:
                    print(f"[OpenFIGI] Batch {i//100+1} failed: {e}, retrying in 5s...")
                    time.sleep(5)
                else:
                    print(f"[OpenFIGI] Batch {i//100+1} failed on retry: {e} — marking {len(batch)} CUSIPs as N/A")
                    for cusip in batch:
                        results[cusip] = "N/A"

        # Rate limit: ~20 req/min = 1 every 3 seconds
        if i + 50 < len(uncached):
            time.sleep(3)

    return results


# ── SEC company_tickers.json name→ticker fallback ─────────────────────────────

_sec_tickers_cache = None
_sec_tickers_lock = threading.Lock()


def _load_sec_tickers():
    """Download SEC company_tickers.json and build normalized name→ticker map."""
    global _sec_tickers_cache
    if _sec_tickers_cache is not None:
        return
    try:
        req = urllib.request.Request(
            "https://www.sec.gov/files/company_tickers.json",
            headers={"User-Agent": "13F-App/1.0 holdings@example.com"},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
        name_map = {}
        for item in data.values():
            norm = item["title"].upper().strip()
            tk = item["ticker"].upper().strip()
            if norm and tk:
                name_map[norm] = tk
        _sec_tickers_cache = name_map
        print(f"[SEC Tickers] Loaded {len(name_map):,} company name→ticker mappings")
    except Exception as e:
        try:
            print(f"[SEC Tickers] Failed to load: {e}")
        except (UnicodeEncodeError, OSError):
            print("[SEC Tickers] Failed to load (encoding error in message)")
        _sec_tickers_cache = {}


def _sec_name_to_ticker(name):
    """Lookup ticker by company name using SEC company_tickers.json."""
    with _sec_tickers_lock:
        _load_sec_tickers()
    if not _sec_tickers_cache or not name:
        return None
    norm = name.upper().strip()
    # Exact match
    if norm in _sec_tickers_cache:
        return _sec_tickers_cache[norm]
    # Try stripping common suffixes
    for suffix in [" INC.", " INC", " CORP.", " CORP", " CORPORATION",
                   " CO.", " CO", " LTD.", " LTD", " PLC", " GROUP",
                   " HOLDINGS", " LP", " N.V.", " SA", " AG", " SE",
                   " & CO.", " & CO", " INTERNATIONAL", " INTL"]:
        stripped = norm.replace(suffix, "").strip()
        if stripped and stripped in _sec_tickers_cache:
            return _sec_tickers_cache[stripped]
    # Try removing "THE " prefix
    if norm.startswith("THE "):
        no_the = norm[4:]
        if no_the in _sec_tickers_cache:
            return _sec_tickers_cache[no_the]
        for suffix in [" INC.", " INC", " CORP.", " CORP", " CORPORATION",
                       " CO.", " CO", " GROUP", " HOLDINGS"]:
            stripped = no_the.replace(suffix, "").strip()
            if stripped and stripped in _sec_tickers_cache:
                return _sec_tickers_cache[stripped]
    return None


def _resolve_remaining_tickers(holdings_list, name_key="name"):
    """Final fallback: resolve any remaining N/A tickers via SEC name lookup."""
    count = 0
    for rec in holdings_list:
        if rec.get("ticker", "N/A") == "N/A":
            found = _sec_name_to_ticker(rec.get(name_key, ""))
            if found:
                rec["ticker"] = found
                count += 1
    if count:
        print(f"[SEC Tickers] Resolved {count} tickers via company name lookup")


# ── Fieldnames ───────────────────────────────────────────────────────────────

FIELDNAMES = [
    # Base fields (8)
    "manager", "period_of_report", "filed_at", "rank",
    "name", "ticker", "value_usd", "pct_of_portfolio",
    # Prior quarter (6)
    "prior_price_qtr_end", "prior_quarter_return_pct",
    "prior_reported_eps", "prior_consensus_eps",
    "prior_eps_beat_dollars", "prior_eps_beat_pct",
    # Filing quarter (6)
    "filing_price_qtr_end", "filing_quarter_return_pct",
    "filing_reported_eps", "filing_consensus_eps",
    "filing_eps_beat_dollars", "filing_eps_beat_pct",
    # Current / live (5)
    "forward_pe", "forward_eps_growth", "dividend_yield",
    "trailing_eps", "forward_eps",
    # QTD (2)
    "qtd_return_pct", "qtd_price_start",
    # Current price (1)
    "current_price",
    # Static (3)
    "sector", "industry", "country",
]

# Original 8 fields (for backward compatibility checks)
BASE_FIELDNAMES = FIELDNAMES[:8]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _safe_name(manager_name):
    return (manager_name.lower()
            .replace(" ", "_").replace("(", "").replace(")", "")
            .replace("/", "").replace("&", "and"))


def _has_enrichment(rows):
    """Check if rows contain enrichment data (any non-None enrichment field)."""
    for r in rows:
        if any(r.get(f) is not None for f in FIELDNAMES[8:]):
            return True
    return False


def _format_xlsx_sheet(ws, rows):
    """Apply professional formatting to a worksheet with holdings data."""
    enriched = _has_enrichment(rows)

    if enriched:
        HEADERS = [
            "Manager", "Period of Report", "Filed At", "Rank",
            "Name", "Ticker", "Value (USD $000s)", "% of Portfolio",
            # Prior quarter
            "Prior Qtr Price", "Prior Qtr Return %",
            "Prior Reported EPS", "Prior Consensus EPS",
            "Prior EPS Beat ($)", "Prior EPS Beat (%)",
            # Filing quarter
            "Filing Qtr Price", "Filing Qtr Return %",
            "Filing Reported EPS", "Filing Consensus EPS",
            "Filing EPS Beat ($)", "Filing EPS Beat (%)",
            # Current / live
            "Fwd P/E (Curr)", "Fwd EPS Growth", "Div Yield",
            "Trail 4Q EPS", "Fwd 12M EPS",
            # QTD
            "QTD Return %",
            # Static
            "Sector", "Industry", "Country",
        ]
    else:
        HEADERS = [
            "Manager", "Period of Report", "Filed At", "Rank",
            "Name", "Ticker", "Value (USD $000s)", "% of Portfolio",
        ]

    num_cols = len(HEADERS)

    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    txt_font = Font(name="Arial", size=10)
    num_font = Font(name="Arial", size=10)
    green_font = Font(name="Arial", size=10, color="227722")
    red_font = Font(name="Arial", size=10, color="CC2222")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"))

    # Headers
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Column indices (1-based) for enriched mode
    # 1-8: base, 9-14: prior qtr, 15-20: filing qtr, 21-25: live, 26: QTD, 27-29: static
    COL_PRIOR_PRICE, COL_PRIOR_RET = 9, 10
    COL_PRIOR_EPS_R, COL_PRIOR_EPS_C = 11, 12
    COL_PRIOR_BEAT_D, COL_PRIOR_BEAT_P = 13, 14
    COL_FILING_PRICE, COL_FILING_RET = 15, 16
    COL_FILING_EPS_R, COL_FILING_EPS_C = 17, 18
    COL_FILING_BEAT_D, COL_FILING_BEAT_P = 19, 20
    COL_FWD_PE, COL_FWD_GROWTH, COL_DIV_YIELD = 21, 22, 23
    COL_TRAIL_EPS, COL_FWD_EPS = 24, 25
    COL_QTD_RET = 26
    COL_SECTOR, COL_INDUSTRY, COL_COUNTRY = 27, 28, 29

    PRICE_COLS = {COL_PRIOR_PRICE, COL_FILING_PRICE}
    RETURN_COLS = {COL_PRIOR_RET, COL_FILING_RET, COL_QTD_RET}
    PE_COLS = {COL_FWD_PE}
    EPS_COLS = {COL_PRIOR_EPS_R, COL_PRIOR_EPS_C, COL_FILING_EPS_R, COL_FILING_EPS_C,
                COL_TRAIL_EPS, COL_FWD_EPS}
    BEAT_D_COLS = {COL_PRIOR_BEAT_D, COL_FILING_BEAT_D}
    BEAT_P_COLS = {COL_PRIOR_BEAT_P, COL_FILING_BEAT_P}
    PCT_COLS = {COL_FWD_GROWTH, COL_DIV_YIELD}
    TEXT_COLS_ENRICHED = {COL_SECTOR, COL_INDUSTRY, COL_COUNTRY}

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        vals = [
            row["manager"],
            row["period_of_report"],
            row["filed_at"],
            int(row["rank"]),
            row["name"],
            row["ticker"],
            int(round(row["value_usd"] / 1000)),
            row["pct_of_portfolio"] / 100,
        ]
        if enriched:
            # Prior quarter
            prior_ret = row.get("prior_quarter_return_pct")
            vals.extend([
                row.get("prior_price_qtr_end"),
                prior_ret / 100 if prior_ret is not None else None,
                row.get("prior_reported_eps"),
                row.get("prior_consensus_eps"),
                row.get("prior_eps_beat_dollars"),
                row.get("prior_eps_beat_pct"),
            ])
            # Filing quarter
            filing_ret = row.get("filing_quarter_return_pct")
            vals.extend([
                row.get("filing_price_qtr_end"),
                filing_ret / 100 if filing_ret is not None else None,
                row.get("filing_reported_eps"),
                row.get("filing_consensus_eps"),
                row.get("filing_eps_beat_dollars"),
                row.get("filing_eps_beat_pct"),
            ])
            # Current / live
            fwd_growth = row.get("forward_eps_growth")
            div_yield = row.get("dividend_yield")
            vals.extend([
                row.get("forward_pe"),
                fwd_growth / 100 if fwd_growth is not None else None,
                div_yield / 100 if div_yield is not None else None,
                row.get("trailing_eps"),
                row.get("forward_eps"),
            ])
            # QTD
            qtd_ret = row.get("qtd_return_pct")
            vals.append(qtd_ret / 100 if qtd_ret is not None else None)
            # Static
            vals.extend([
                row.get("sector") or "",
                row.get("industry") or "",
                row.get("country") or "",
            ])

        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

            # Base text columns (A-C, E-F)
            if c_idx in (1, 2, 3, 5, 6):
                cell.font = txt_font
                cell.number_format = "@"
            # Rank (D)
            elif c_idx == 4:
                cell.font = num_font
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            # Value (G)
            elif c_idx == 7:
                cell.font = num_font
                cell.number_format = "#,##0"
            # Pct of portfolio (H)
            elif c_idx == 8:
                cell.font = num_font
                cell.number_format = "0.00%"
            elif not enriched:
                continue
            # Price columns
            elif c_idx in PRICE_COLS:
                cell.font = num_font
                cell.number_format = "$#,##0.00"
            # Return columns — percentage with color
            elif c_idx in RETURN_COLS:
                if val is not None:
                    cell.font = green_font if val >= 0 else red_font
                    cell.number_format = "0.00%"
                else:
                    cell.font = txt_font
            # P/E columns — 0.0x format
            elif c_idx in PE_COLS:
                cell.font = num_font
                cell.number_format = '0.0"x"'
            # EPS columns
            elif c_idx in EPS_COLS:
                cell.font = num_font
                cell.number_format = "0.00"
            # EPS Beat $ columns — green/red
            elif c_idx in BEAT_D_COLS:
                if val is not None:
                    cell.font = green_font if val >= 0 else red_font
                    cell.number_format = '+0.00;-0.00'
                else:
                    cell.font = txt_font
            # EPS Beat % columns — green/red
            elif c_idx in BEAT_P_COLS:
                if val is not None:
                    cell.font = green_font if val >= 0 else red_font
                    cell.number_format = '+0.0%;-0.0%'
                else:
                    cell.font = txt_font
            # Growth / Yield columns
            elif c_idx in PCT_COLS:
                cell.font = num_font
                cell.number_format = "0.0%"
            # Sector, Industry
            elif c_idx in TEXT_COLS_ENRICHED:
                cell.font = txt_font

    # Alternating row fill
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    for r_idx in range(3, len(rows) + 2, 2):
        for c_idx in range(1, num_cols + 1):
            ws.cell(row=r_idx, column=c_idx).fill = alt_fill

    # Column widths
    if enriched:
        widths = [
            22, 16, 14, 6, 30, 8, 18, 14,      # base (8)
            12, 12, 11, 12, 12, 12, 12,          # prior qtr (7)
            12, 12, 11, 12, 12, 12, 12,          # filing qtr (7)
            11, 12, 10,                           # live (3)
            18, 22, 18,                           # static (3)
        ]
    else:
        widths = [22, 16, 14, 6, 30, 8, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _eps_beat_label(beat_dollars, beat_pct=None):
    """Format EPS beat as '+$0.12 (+8.5%)' or 'N/A'."""
    if beat_dollars is None:
        return "N/A"
    sign = "+" if beat_dollars >= 0 else ""
    dollar_part = f"{sign}${beat_dollars:.2f}"
    if beat_pct is not None:
        pct_part = f"({sign}{beat_pct:.1f}%)"
        return f"{dollar_part} {pct_part}"
    return dollar_part


def write_individual_xlsx(manager_name, period, rows):
    """Write a per-manager Excel file and return the filename."""
    safe = _safe_name(manager_name)
    period_str = period.replace("-", "")
    filename = f"{safe}_top20_{period_str}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = manager_name[:31]
    _format_xlsx_sheet(ws, rows)
    wb.save(path)
    return filename


def write_combined_xlsx(all_rows, run_date):
    """Write combined Excel for all managers and return the filename."""
    filename = f"all_managers_top20_{run_date}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "All Managers"
    _format_xlsx_sheet(ws, all_rows)
    wb.save(path)
    return filename


def write_weighted_xlsx(all_rows, run_date, manager_weights=None):
    """
    Write a weighted combined portfolio Excel that deduplicates stocks across
    managers and weights them by the user's manager allocations.

    Returns the filename.
    """
    import analysis

    weighted_rows, _ = analysis._apply_manager_weights(all_rows, manager_weights)

    # Aggregate by ticker
    by_ticker = {}
    for r in weighted_rows:
        tk = r.get("ticker", "N/A")
        key = tk if tk != "N/A" else r.get("name", "Unknown")
        if key not in by_ticker:
            by_ticker[key] = {
                "ticker": r.get("ticker", "N/A"),
                "name": r.get("name", "Unknown"),
                "combined_weight": 0.0,
                "total_value": 0,
                "managers": set(),
                # Enrichment fields (per-ticker, take from first non-None)
                "sector": None,
                "industry": None,
                "country": None,
                "prior_quarter_return_pct": None,
                "filing_quarter_return_pct": None,
                "forward_pe": None,
                "forward_eps_growth": None,
                "dividend_yield": None,
                "trailing_eps": None,
                "forward_eps": None,
                "qtd_return_pct": None,
                "prior_price_qtr_end": None,
                "filing_price_qtr_end": None,
                "prior_reported_eps": None,
                "prior_consensus_eps": None,
                "prior_eps_beat_dollars": None,
                "prior_eps_beat_pct": None,
                "filing_reported_eps": None,
                "filing_consensus_eps": None,
                "filing_eps_beat_dollars": None,
                "filing_eps_beat_pct": None,
            }
        d = by_ticker[key]
        d["combined_weight"] += r.get("combined_weight", 0)
        d["total_value"] += r.get("value_usd", 0)
        d["managers"].add(r["manager"])
        # Fill enrichment from first row that has data (all same per ticker)
        for field in [
            "sector", "industry", "country",
            "prior_quarter_return_pct", "filing_quarter_return_pct",
            "forward_pe", "forward_eps_growth", "dividend_yield",
            "trailing_eps", "forward_eps", "qtd_return_pct",
            "prior_price_qtr_end", "filing_price_qtr_end",
            "prior_reported_eps", "prior_consensus_eps",
            "prior_eps_beat_dollars", "prior_eps_beat_pct",
            "filing_reported_eps", "filing_consensus_eps",
            "filing_eps_beat_dollars", "filing_eps_beat_pct",
        ]:
            if d[field] is None and r.get(field) is not None:
                d[field] = r[field]

    # Sort by combined weight descending
    sorted_stocks = sorted(by_ticker.values(), key=lambda x: -x["combined_weight"])

    # Build workbook
    filename = f"weighted_portfolio_{run_date}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "Weighted Portfolio"

    enriched = any(s["sector"] is not None or s["industry"] is not None for s in sorted_stocks)

    if enriched:
        HEADERS = [
            "Rank", "Ticker", "Name", "Sector", "Industry", "Country",
            "Wtd Port %", "# Managers", "Managers", "Total Value ($000s)",
            "Prior Qtr Price", "Prior Qtr Return %",
            "Prior Reported EPS", "Prior Consensus EPS",
            "Prior EPS Beat ($)", "Prior EPS Beat (%)",
            "Filing Qtr Price", "Filing Qtr Return %",
            "Filing Reported EPS", "Filing Consensus EPS",
            "Filing EPS Beat ($)", "Filing EPS Beat (%)",
            "Fwd P/E", "Fwd EPS Growth %", "Div Yield %",
            "Trail 4Q EPS", "Fwd 12M EPS",
            "QTD Return %",
        ]
    else:
        HEADERS = [
            "Rank", "Ticker", "Name",
            "Wtd Port %", "# Managers", "Managers", "Total Value ($000s)",
        ]

    num_cols = len(HEADERS)

    # Styles
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    txt_font = Font(name="Arial", size=10)
    num_font = Font(name="Arial", size=10)
    green_font = Font(name="Arial", size=10, color="227722")
    red_font = Font(name="Arial", size=10, color="CC2222")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"))

    # Write headers
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Write data rows
    for r_idx, s in enumerate(sorted_stocks, 2):
        rank = r_idx - 1
        mgr_list = ", ".join(sorted(s["managers"]))
        mgr_count = len(s["managers"])
        wt_pct = s["combined_weight"] / 100  # as decimal for Excel % format

        if enriched:
            prior_ret = s["prior_quarter_return_pct"]
            filing_ret = s["filing_quarter_return_pct"]
            fwd_growth = s["forward_eps_growth"]
            div_yield = s["dividend_yield"]
            qtd_ret = s.get("qtd_return_pct")
            vals = [
                rank, s["ticker"], s["name"],
                s["sector"] or "", s["industry"] or "", s["country"] or "",
                wt_pct, mgr_count, mgr_list,
                int(round(s["total_value"] / 1000)) if s["total_value"] else 0,
                s["prior_price_qtr_end"],
                prior_ret / 100 if prior_ret is not None else None,
                s["prior_reported_eps"], s["prior_consensus_eps"],
                s["prior_eps_beat_dollars"], s["prior_eps_beat_pct"],
                s["filing_price_qtr_end"],
                filing_ret / 100 if filing_ret is not None else None,
                s["filing_reported_eps"], s["filing_consensus_eps"],
                s["filing_eps_beat_dollars"], s["filing_eps_beat_pct"],
                s["forward_pe"],
                fwd_growth / 100 if fwd_growth is not None else None,
                div_yield / 100 if div_yield is not None else None,
                s["trailing_eps"],
                s["forward_eps"],
                qtd_ret / 100 if qtd_ret is not None else None,
            ]
        else:
            vals = [
                rank, s["ticker"], s["name"],
                wt_pct, mgr_count, mgr_list,
                int(round(s["total_value"] / 1000)) if s["total_value"] else 0,
            ]

        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

            if enriched:
                # Column mapping for enriched mode:
                # 1=Rank, 2=Ticker, 3=Name, 4=Sector, 5=Industry, 6=Country,
                # 7=Wtd%, 8=#Mgrs, 9=Managers, 10=Value,
                # 11=PriorPrice, 12=PriorRet,
                # 13=PriorEPSR, 14=PriorEPSC, 15=PriorBeat$, 16=PriorBeat%,
                # 17=FilingPrice, 18=FilingRet,
                # 19=FilingEPSR, 20=FilingEPSC, 21=FilingBeat$, 22=FilingBeat%,
                # 23=FwdPE, 24=FwdGrowth, 25=DivYield, 26=TrailEPS, 27=FwdEPS,
                # 28=QTDRet
                if c_idx == 1:  # Rank
                    cell.font = num_font
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx in (2, 3, 4, 5, 6, 9):  # Text cols
                    cell.font = txt_font
                elif c_idx == 7:  # Wtd Port %
                    cell.font = Font(name="Arial", size=10, bold=True)
                    cell.number_format = "0.00%"
                elif c_idx == 8:  # # Managers
                    cell.font = num_font
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx == 10:  # Total Value
                    cell.font = num_font
                    cell.number_format = "#,##0"
                elif c_idx in (11, 17):  # Price cols
                    cell.font = num_font
                    cell.number_format = "$#,##0.00"
                elif c_idx in (12, 18, 28):  # Return % cols (prior, filing, QTD)
                    if val is not None:
                        cell.font = green_font if val >= 0 else red_font
                        cell.number_format = "0.00%"
                    else:
                        cell.font = txt_font
                elif c_idx == 23:  # Fwd P/E
                    cell.font = num_font
                    cell.number_format = '0.0"x"'
                elif c_idx in (13, 14, 19, 20, 26, 27):  # EPS cols (incl. Trail/Fwd EPS)
                    cell.font = num_font
                    cell.number_format = "0.00"
                elif c_idx in (15, 21):  # Beat $ cols
                    if val is not None:
                        cell.font = green_font if val >= 0 else red_font
                        cell.number_format = '+0.00;-0.00'
                    else:
                        cell.font = txt_font
                elif c_idx in (16, 22):  # Beat % cols
                    if val is not None:
                        cell.font = green_font if val >= 0 else red_font
                        cell.number_format = '+0.0%;-0.0%'
                    else:
                        cell.font = txt_font
                elif c_idx in (24, 25):  # Growth / Yield
                    cell.font = num_font
                    cell.number_format = "0.0%"
            else:
                # Non-enriched column mapping:
                # 1=Rank, 2=Ticker, 3=Name, 4=Wtd%, 5=#Mgrs, 6=Managers, 7=Value
                if c_idx == 1:
                    cell.font = num_font
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx in (2, 3, 6):
                    cell.font = txt_font
                elif c_idx == 4:
                    cell.font = Font(name="Arial", size=10, bold=True)
                    cell.number_format = "0.00%"
                elif c_idx == 5:
                    cell.font = num_font
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx == 7:
                    cell.font = num_font
                    cell.number_format = "#,##0"

    # Alternating row fill
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    for r_idx in range(3, len(sorted_stocks) + 2, 2):
        for c_idx in range(1, num_cols + 1):
            ws.cell(row=r_idx, column=c_idx).fill = alt_fill

    # Column widths
    if enriched:
        widths = [
            6, 8, 30, 18, 22, 18,                    # Rank..Country
            12, 10, 40, 18,                           # Wtd%..Value
            12, 12, 11, 12, 12, 12, 12,              # Prior qtr (7)
            12, 12, 11, 12, 12, 12, 12,              # Filing qtr (7)
            11, 12, 10,                               # Fwd PE, Growth, Yield
        ]
    else:
        widths = [6, 8, 30, 12, 10, 40, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}{len(sorted_stocks) + 1}"

    wb.save(path)
    return filename


# ── Simplified Excel (single-file export) ──────────────────────────────────

def write_simplified_xlsx(all_rows, run_date, manager_weights=None):
    """
    Write a single-file Excel export with:
      - Sheet 1: "Weighted Portfolio" with title noting manager weightings
      - Sheet 2+: One sheet per individual manager
    ALL holdings shown, ranked by % of portfolio. 12 columns matching UI tables.
    Returns the filename.
    """
    import analysis

    weighted_rows, _ = analysis._apply_manager_weights(all_rows, manager_weights)

    # ── Styles ──
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    txt_font = Font(name="Arial", size=10)
    num_font = Font(name="Arial", size=10)
    green_font = Font(name="Arial", size=10, color="227722")
    red_font = Font(name="Arial", size=10, color="CC2222")
    yellow_font = Font(name="Arial", size=10, color="AA8800")
    bold_font = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=12, bold=True, color="2F5496")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"))
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    totals_fill = PatternFill("solid", fgColor="E8EDF5")

    HEADERS = [
        "% of Port", "Stock Name", "Ticker", "Sector", "Industry",
        "Qtr End Price", "Current Price", "QTD Return",
        "Fwd P/E", "Fwd EPS Growth",
        "Reported EPS", "EPS Beat",
    ]

    def _write_sheet(ws, stocks, sheet_title=None, weight_note=None):
        """Write one sheet of holdings data with totals row."""
        start_row = 1
        if sheet_title or weight_note:
            ws.cell(row=1, column=1, value=sheet_title or "").font = title_font
            if weight_note:
                ws.cell(row=2, column=1, value=weight_note).font = Font(name="Arial", size=9, color="666666")
            start_row = 4

        # Headers
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        # Data rows
        for r_idx, s in enumerate(stocks, start_row + 1):
            pct = s.get("pct", 0)
            qtd = s.get("qtd_return")
            fwd_pe = s.get("forward_pe")
            fwd_growth = s.get("forward_eps_growth")
            beat_d = s.get("eps_beat_dollars")
            beat_p = s.get("eps_beat_pct")

            # EPS beat indicator
            if beat_d is not None:
                if beat_d > 0:
                    beat_str = f"\u2191 +{beat_p:.0f}%" if beat_p is not None else "\u2191 Beat"
                elif beat_d < 0:
                    beat_str = f"\u2193 {beat_p:.0f}%" if beat_p is not None else "\u2193 Miss"
                else:
                    beat_str = "\u2192 Met"
            else:
                beat_str = ""

            vals = [
                pct / 100 if pct else 0,  # % as decimal for Excel
                s.get("name", ""),
                s.get("ticker", ""),
                s.get("sector", ""),
                s.get("industry", ""),
                s.get("filing_price"),
                s.get("current_price"),
                qtd / 100 if qtd is not None else None,
                fwd_pe,
                fwd_growth / 100 if fwd_growth is not None else None,
                s.get("filing_reported_eps"),
                beat_str,
            ]

            # Append monthly return columns if available
            for m in s.get("monthly_returns", []):
                mr = m.get("return_pct")
                vals.append(mr / 100 if mr is not None else None)

            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if c_idx == 1:  # % of Port
                    cell.font = bold_font
                    cell.number_format = "0.00%"
                elif c_idx in (2, 3, 4, 5):  # Name, Ticker, Sector, Industry
                    cell.font = txt_font
                elif c_idx in (6, 7):  # Prices
                    cell.font = num_font
                    cell.number_format = "$#,##0.00"
                elif c_idx == 8:  # QTD Return
                    if val is not None:
                        cell.font = green_font if val >= 0 else red_font
                        cell.number_format = "0.00%"
                elif c_idx == 9:  # Fwd P/E
                    cell.font = num_font
                    cell.number_format = '0.0"x"'
                elif c_idx == 10:  # Fwd EPS Growth
                    cell.font = num_font
                    cell.number_format = "0.0%"
                elif c_idx == 11:  # Reported EPS
                    cell.font = num_font
                    cell.number_format = "0.00"
                elif c_idx == 12:  # EPS Beat indicator
                    if beat_d is not None:
                        cell.font = green_font if beat_d > 0 else (red_font if beat_d < 0 else yellow_font)
                elif c_idx > 12:  # Monthly return columns
                    if val is not None:
                        cell.font = green_font if val >= 0 else red_font
                        cell.number_format = "0.00%"

            # Alternating rows
            if (r_idx - start_row) % 2 == 0:
                for c_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = alt_fill

        # Totals row
        totals_row = start_row + 1 + len(stocks)
        # Compute weighted totals from ALL stocks
        total_pct = sum(s.get("pct", 0) for s in stocks)
        qtd_sum, qtd_wt = 0.0, 0.0
        pe_inv, pe_wt = 0.0, 0.0
        eg_sum, eg_wt = 0.0, 0.0
        for s in stocks:
            w = s.get("pct", 0)
            if w <= 0:
                continue
            if s.get("qtd_return") is not None:
                qtd_sum += s["qtd_return"] * w
                qtd_wt += w
            if s.get("forward_pe") is not None and s["forward_pe"] > 0:
                pe_inv += w / s["forward_pe"]
                pe_wt += w
            if s.get("forward_eps_growth") is not None:
                clamped = max(-50, min(50, s["forward_eps_growth"]))
                eg_sum += clamped * w
                eg_wt += w

        # Determine total columns including monthly returns
        n_monthly = len(stocks[0].get("monthly_returns", [])) if stocks else 0
        total_cols = len(HEADERS) + n_monthly

        ws.cell(row=totals_row, column=1, value=total_pct / 100 if total_pct else 0)
        ws.cell(row=totals_row, column=1).font = bold_font
        ws.cell(row=totals_row, column=1).number_format = "0.00%"
        ws.cell(row=totals_row, column=2, value="TOTAL (Weighted)").font = bold_font
        if qtd_wt > 0:
            wtd_qtd = qtd_sum / qtd_wt / 100
            cell_qtd = ws.cell(row=totals_row, column=8, value=wtd_qtd)
            cell_qtd.font = green_font if wtd_qtd >= 0 else red_font
            cell_qtd.number_format = "0.00%"
        if pe_inv > 0:
            ws.cell(row=totals_row, column=9, value=pe_wt / pe_inv).font = bold_font
            ws.cell(row=totals_row, column=9).number_format = '0.0"x"'
        if eg_wt > 0:
            ws.cell(row=totals_row, column=10, value=eg_sum / eg_wt / 100).font = bold_font
            ws.cell(row=totals_row, column=10).number_format = "0.0%"
        for c_idx in range(1, total_cols + 1):
            ws.cell(row=totals_row, column=c_idx).border = thin_border
            ws.cell(row=totals_row, column=c_idx).fill = totals_fill

        # Column widths
        widths = [10, 28, 8, 18, 22, 12, 12, 12, 10, 12, 12, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Monthly return column widths
        for i in range(n_monthly):
            ws.column_dimensions[get_column_letter(len(HEADERS) + i + 1)].width = 10

        # Write monthly return headers if present
        if n_monthly and stocks:
            for mi, m in enumerate(stocks[0].get("monthly_returns", [])):
                col = len(HEADERS) + mi + 1
                cell = ws.cell(row=start_row, column=col, value=m.get("month", ""))
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = thin_border

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
        last_col = get_column_letter(total_cols)
        ws.auto_filter.ref = f"A{start_row}:{last_col}{totals_row}"

        # Methodology footnote
        fn_row = totals_row + 2
        fn_font = Font(size=8, italic=True, color="808080")
        ws.cell(row=fn_row, column=1, value="Methodology:").font = Font(size=8, bold=True, italic=True, color="808080")
        ws.cell(row=fn_row + 1, column=1,
                value="Fwd EPS Growth: Analyst consensus next-fiscal-year EPS growth estimate (source: Yahoo Finance growth_estimates). "
                      "Portfolio-level growth is winsorized at \u00b150% per stock before weighting to limit outlier impact.").font = fn_font
        ws.merge_cells(start_row=fn_row + 1, start_column=1, end_row=fn_row + 1, end_column=total_cols)
        ws.cell(row=fn_row + 2, column=1,
                value="Fwd P/E: Forward price/earnings (source: Yahoo Finance forwardPE = currentPrice / forwardEps). "
                      "Portfolio P/E is weighted harmonic mean; stocks with negative values are excluded.").font = fn_font
        ws.merge_cells(start_row=fn_row + 2, start_column=1, end_row=fn_row + 2, end_column=total_cols)
        ws.cell(row=fn_row + 3, column=1,
                value="QTD Return: Weighted arithmetic mean. Monthly returns computed from yfinance historical prices. "
                      "All totals computed from full holdings, not just displayed rows.").font = fn_font
        ws.merge_cells(start_row=fn_row + 3, start_column=1, end_row=fn_row + 3, end_column=total_cols)

    def _stocks_from_weighted(rows):
        """Aggregate weighted rows by ticker and sort by combined_weight desc."""
        by_ticker = {}
        for r in rows:
            tk = r.get("ticker", "N/A")
            key = tk if tk != "N/A" else r.get("name", "Unknown")
            if key not in by_ticker:
                by_ticker[key] = {
                    "pct": 0, "name": analysis.shorten_stock_name(r.get("name", "")),
                    "ticker": tk, "sector": None, "industry": None,
                    "filing_price": None, "current_price": None,
                    "qtd_return": None, "forward_pe": None, "forward_eps_growth": None,
                    "filing_reported_eps": None, "eps_beat_dollars": None,
                    "eps_beat_pct": None, "monthly_returns": None,
                }
            d = by_ticker[key]
            d["pct"] += r.get("combined_weight", r.get("pct_of_portfolio", 0))
            if d["sector"] is None and r.get("sector"):
                d["sector"] = r["sector"]
            if d["industry"] is None and r.get("industry"):
                d["industry"] = r["industry"]
            if d["monthly_returns"] is None and r.get("monthly_returns"):
                d["monthly_returns"] = r["monthly_returns"]
            for field, src in [
                ("filing_price", "filing_price_qtr_end"), ("current_price", "current_price"),
                ("qtd_return", "qtd_return_pct"), ("forward_pe", "forward_pe"),
                ("forward_eps_growth", "forward_eps_growth"),
                ("filing_reported_eps", "filing_reported_eps"),
                ("eps_beat_dollars", "filing_eps_beat_dollars"),
                ("eps_beat_pct", "filing_eps_beat_pct"),
            ]:
                if d[field] is None and r.get(src) is not None:
                    d[field] = r[src]
        return sorted(by_ticker.values(), key=lambda x: -x["pct"])

    def _stocks_from_manager(rows):
        """Build stock list from a single manager's rows."""
        stocks = []
        for r in sorted(rows, key=lambda x: -x.get("pct_of_portfolio", 0)):
            stocks.append({
                "pct": r.get("pct_of_portfolio", 0),
                "name": analysis.shorten_stock_name(r.get("name", "")),
                "ticker": r.get("ticker", "N/A"),
                "sector": r.get("sector", ""),
                "industry": r.get("industry", ""),
                "filing_price": r.get("filing_price_qtr_end"),
                "current_price": r.get("current_price"),
                "qtd_return": r.get("qtd_return_pct"),
                "forward_pe": r.get("forward_pe"),
                "forward_eps_growth": r.get("forward_eps_growth"),
                "filing_reported_eps": r.get("filing_reported_eps"),
                "eps_beat_dollars": r.get("filing_eps_beat_dollars"),
                "eps_beat_pct": r.get("filing_eps_beat_pct"),
                "monthly_returns": r.get("monthly_returns"),
            })
        return stocks

    # Build workbook
    filename = f"portfolio_{run_date}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb = Workbook()

    # Sheet 1: Weighted Portfolio
    ws_w = wb.active
    ws_w.title = "Weighted Portfolio"
    managers = list({r["manager"] for r in all_rows})
    wts = manager_weights or {}
    wt_parts = []
    for m in sorted(managers):
        w = wts.get(m)
        wt_parts.append(f"{m}: {w:.0f}%" if w else f"{m}: equal")
    weight_note = "Weightings: " + ", ".join(wt_parts) if wt_parts else ""
    weighted_stocks = _stocks_from_weighted(weighted_rows)
    _write_sheet(ws_w, weighted_stocks, sheet_title="Weighted Portfolio", weight_note=weight_note)

    # Per-manager sheets
    by_mgr = defaultdict(list)
    for r in all_rows:
        by_mgr[r["manager"]].append(r)
    for mgr in sorted(by_mgr.keys()):
        safe_name = mgr[:31].replace("/", "-").replace("\\", "-")
        ws_m = wb.create_sheet(title=safe_name)
        mgr_stocks = _stocks_from_manager(by_mgr[mgr])
        _write_sheet(ws_m, mgr_stocks, sheet_title=mgr)

    wb.save(path)
    return filename


# ── Legacy CSV support ───────────────────────────────────────────────────────

def write_individual_csv(manager_name, period, rows):
    """Write a per-manager CSV (legacy). Prefer write_individual_xlsx."""
    safe = _safe_name(manager_name)
    period_str = period.replace("-", "")
    filename = f"{safe}_top20_{period_str}.csv"
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return filename


# ── Row builder ──────────────────────────────────────────────────────────────

def _build_rows(manager_name, period, filed_at, holdings_list, value_key="value", multiplier=1):
    """
    Convert a list of (name, ticker, value) dicts into ranked rows.
    holdings_list: list of dicts with keys 'name', 'ticker', value_key
    """
    total_value = sum(h.get(value_key, 0) or 0 for h in holdings_list) * multiplier
    if total_value == 0:
        total_value = 1

    rows = []
    for rank, h in enumerate(holdings_list[:TOP_N], start=1):
        val = (h.get(value_key, 0) or 0) * multiplier
        pct = round(val / total_value * 100, 2)
        rows.append({
            "manager": manager_name,
            "period_of_report": period,
            "filed_at": filed_at,
            "rank": rank,
            "name": h.get("name", "Unknown"),
            "ticker": h.get("ticker", "N/A"),
            "value_usd": round(val),
            "pct_of_portfolio": pct,
            # Prior quarter enrichment (6)
            "prior_price_qtr_end": None,
            "prior_quarter_return_pct": None,
            "prior_reported_eps": None,
            "prior_consensus_eps": None,
            "prior_eps_beat_dollars": None,
            "prior_eps_beat_pct": None,
            # Filing quarter enrichment (6)
            "filing_price_qtr_end": None,
            "filing_quarter_return_pct": None,
            "filing_reported_eps": None,
            "filing_consensus_eps": None,
            "filing_eps_beat_dollars": None,
            "filing_eps_beat_pct": None,
            # Current / live (5)
            "forward_pe": None,
            "forward_eps_growth": None,
            "dividend_yield": None,
            "trailing_eps": None,
            "forward_eps": None,
            # QTD (2)
            "qtd_return_pct": None,
            "qtd_price_start": None,
            # Current price (1)
            "current_price": None,
            # Static (3)
            "sector": None,
            "industry": None,
            "country": None,
        })
    return rows


# ── 13F Fetcher ──────────────────────────────────────────────────────────────

def fetch_13f(manager_name, cik):
    """
    Fetch latest 13F-HR filing for a given CIK.
    Returns (period, filed_at, n_total, rows).
    """
    company = Company(cik)
    filings = company.get_filings(form="13F-HR")
    if not filings or len(filings) == 0:
        raise ValueError(f"No 13F-HR filings found for CIK {cik}")

    filing = None
    obj = None
    cutoff = _max_filing_date()
    for f in filings:
        if str(f.filing_date) > cutoff:
            continue
        _obj = f.obj()
        _por = str(_obj.period_of_report) if hasattr(_obj, "period_of_report") else str(f.filing_date)
        if _por > MAX_DATE:
            continue  # report period is after our target quarter
        filing = f
        obj = _obj
        break

    if filing is None:
        raise ValueError(f"No 13F-HR filing found before {cutoff} with period <= {MAX_DATE}")

    period = str(filing.filing_date)
    filed_at = str(filing.filing_date)

    holdings_df = None
    if hasattr(obj, "infotable"):
        holdings_df = obj.infotable
    elif hasattr(obj, "holdings"):
        holdings_df = obj.holdings

    if holdings_df is None or len(holdings_df) == 0:
        raise ValueError("Could not parse 13F holdings table")

    if hasattr(obj, "period_of_report"):
        period = str(obj.period_of_report)

    cols = {c.lower().replace(" ", "").replace("_", ""): c for c in holdings_df.columns}

    def _find_col(*candidates):
        for cand in candidates:
            key = cand.lower().replace(" ", "").replace("_", "")
            if key in cols:
                return cols[key]
        return None

    col_name  = _find_col("nameOfIssuer", "Name of Issuer", "Issuer", "name", "companyName")
    col_value = _find_col("value", "Value", "value(x$1000)", "marketValue", "mktVal")
    col_cusip = _find_col("cusip", "CUSIP")
    col_tick  = _find_col("ticker", "Ticker", "symbol", "Symbol")

    records = []
    n_total = len(holdings_df)

    for _, row in holdings_df.iterrows():
        name  = str(row[col_name]) if col_name and col_name in row.index else "Unknown"
        cusip = str(row[col_cusip]).strip() if col_cusip and col_cusip in row.index else ""
        value = float(row[col_value] or 0) if col_value and col_value in row.index else 0.0

        ticker = "N/A"
        if col_tick and col_tick in row.index:
            tv = str(row[col_tick]).strip()
            if tv and tv not in ("N/A", "nan", "None", ""):
                ticker = tv
        if ticker == "N/A" and cusip:
            ticker = CUSIP_TO_TICKER.get(cusip, "N/A")

        records.append({"name": name, "ticker": ticker, "value": value, "_cusip": cusip})

    # Batch resolve unresolved CUSIPs via OpenFIGI
    unresolved = {}
    for idx, rec in enumerate(records):
        if rec["ticker"] == "N/A" and rec.get("_cusip"):
            unresolved.setdefault(rec["_cusip"], []).append(idx)
    if unresolved:
        try:
            figi = openfigi_lookup(list(unresolved.keys()))
            for cusip, ticker in figi.items():
                if ticker != "N/A" and cusip in unresolved:
                    for idx in unresolved[cusip]:
                        records[idx]["ticker"] = ticker
        except Exception:
            pass

    _apply_ticker_aliases(records)
    _resolve_remaining_tickers(records)
    records.sort(key=lambda x: x["value"], reverse=True)
    rows = _build_rows(manager_name, period, filed_at, records, value_key="value", multiplier=1)
    return period, filed_at, n_total, rows


# ── N-PORT Fetcher ───────────────────────────────────────────────────────────

def fetch_nport(manager_name, cik, series_keyword):
    """
    Fetch latest NPORT-P filing for a given CIK, matching a series by keyword.
    Returns (period, filed_at, n_total, rows).
    """
    company = Company(cik)
    filings = company.get_filings(form="NPORT-P")
    if not filings or len(filings) == 0:
        raise ValueError(f"No NPORT-P filings found for CIK {cik}")

    ns = {"nport": "http://www.sec.gov/edgar/nport"}

    cutoff = _max_filing_date()
    for f in filings:
        if str(f.filing_date) > cutoff:
            continue

        xml_content = None
        for att in f.attachments:
            if hasattr(att, "is_xml") and att.is_xml:
                xml_content = att.download()
                break
            elif hasattr(att, "document") and att.document.endswith(".xml"):
                xml_content = att.download()
                break

        if not xml_content:
            continue

        if isinstance(xml_content, bytes):
            xml_content = xml_content.decode("utf-8", errors="replace")

        try:
            root = ET.fromstring(xml_content)
        except ET.ParseError:
            continue

        series_el = root.find(".//nport:seriesName", ns)
        if series_el is None:
            series_el = root.find(".//{http://www.sec.gov/edgar/nport}seriesName")
        if series_el is None:
            series_el = root.find(".//seriesName")

        series_name = ""
        if series_el is not None and series_el.text:
            series_name = series_el.text.strip()

        if series_keyword.lower() not in series_name.lower():
            continue

        period = str(f.filing_date)
        filed_at = str(f.filing_date)

        rep_date_el = root.find(".//nport:repPdDate", ns)
        if rep_date_el is not None and rep_date_el.text:
            period = rep_date_el.text.strip()

        # Skip filings whose reporting period is after MAX_DATE
        if period > MAX_DATE:
            continue

        inv_elements = root.findall(".//nport:invstOrSec", ns)
        if not inv_elements:
            inv_elements = root.findall(".//{http://www.sec.gov/edgar/nport}invstOrSec")

        holdings_list = []
        for inv in inv_elements:
            name_el = inv.find("nport:name", ns)
            if name_el is None:
                name_el = inv.find("{http://www.sec.gov/edgar/nport}name")
            name = name_el.text.strip() if name_el is not None and name_el.text else "Unknown"

            val_el = inv.find("nport:valUSD", ns)
            if val_el is None:
                val_el = inv.find("{http://www.sec.gov/edgar/nport}valUSD")
            value = float(val_el.text) if val_el is not None and val_el.text else 0

            cusip_el = inv.find("nport:cusip", ns)
            if cusip_el is None:
                cusip_el = inv.find("{http://www.sec.gov/edgar/nport}cusip")
            cusip = cusip_el.text.strip() if cusip_el is not None and cusip_el.text else ""

            ticker = "N/A"

            ident = inv.find(".//nport:identifiers", ns)
            if ident is None:
                ident = inv.find(".//{http://www.sec.gov/edgar/nport}identifiers")
            if ident is not None:
                ticker_el = ident.find("nport:ticker", ns)
                if ticker_el is None:
                    ticker_el = ident.find("{http://www.sec.gov/edgar/nport}ticker")
                if ticker_el is not None:
                    tv = ticker_el.get("value", "").strip()
                    if tv and tv != "N/A":
                        ticker = tv

            if ticker == "N/A":
                ticker_el2 = inv.find("nport:ticker", ns)
                if ticker_el2 is None:
                    ticker_el2 = inv.find("{http://www.sec.gov/edgar/nport}ticker")
                if ticker_el2 is not None and ticker_el2.text:
                    tv2 = ticker_el2.text.strip()
                    if tv2 and tv2 != "N/A":
                        ticker = tv2

            if ticker == "N/A" and cusip:
                ticker = CUSIP_TO_TICKER.get(cusip, "N/A")

            holdings_list.append({"name": name, "ticker": ticker, "value": value, "_cusip": cusip})

        n_total = len(holdings_list)
        if n_total == 0:
            continue

        # Batch resolve unresolved CUSIPs via OpenFIGI
        unresolved = {}
        for idx, rec in enumerate(holdings_list):
            if rec["ticker"] == "N/A" and rec.get("_cusip"):
                unresolved.setdefault(rec["_cusip"], []).append(idx)
        if unresolved:
            try:
                figi = openfigi_lookup(list(unresolved.keys()))
                for cusip_k, ticker_v in figi.items():
                    if ticker_v != "N/A" and cusip_k in unresolved:
                        for idx in unresolved[cusip_k]:
                            holdings_list[idx]["ticker"] = ticker_v
            except Exception:
                pass

        _apply_ticker_aliases(holdings_list)
        _resolve_remaining_tickers(holdings_list)
        holdings_list.sort(key=lambda x: x["value"], reverse=True)

        rows = _build_rows(manager_name, period, filed_at, holdings_list,
                           value_key="value", multiplier=1)
        return period, filed_at, n_total, rows

    raise ValueError(f"No NPORT-P filing matched series keyword '{series_keyword}' before {cutoff} (max_date={MAX_DATE})")


# ── Main (standalone) ────────────────────────────────────────────────────────

def main():
    combined_rows = []
    run_date = datetime.today().strftime("%Y%m%d")

    print("=" * 60)
    print("Investment Manager Holdings Fetcher — 13F + N-PORT (edgartools)")
    print(f"Max period: {MAX_DATE}  |  Run: {run_date}")
    print("=" * 60)

    for manager_name, cik in MANAGERS_13F.items():
        print(f"\n[13F] {manager_name}  (CIK: {cik})")
        try:
            period, filed_at, n_total, rows = fetch_13f(manager_name, cik)
            print(f"      Period: {period}  |  Filed: {filed_at}  |  Total positions: {n_total}")
            filename = write_individual_csv(manager_name, period, rows)
            combined_rows.extend(rows)
            print(f"      OK {filename}")
        except Exception as e:
            print(f"      Error: {e}")

    for manager_name, info in MANAGERS_NPORT.items():
        print(f"\n[N-PORT] {manager_name}  (CIK: {info['cik']})")
        try:
            period, filed_at, n_total, rows = fetch_nport(
                manager_name, info["cik"], info["series_keyword"]
            )
            print(f"         Period: {period}  |  Filed: {filed_at}  |  Total holdings: {n_total}")
            filename = write_individual_csv(manager_name, period, rows)
            combined_rows.extend(rows)
            print(f"         OK {filename}")
        except Exception as e:
            print(f"         Error: {e}")

    if combined_rows:
        combined_name = f"all_managers_top20_{run_date}.csv"
        combined_path = os.path.join(OUTPUT_DIR, combined_name)
        with open(combined_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(combined_rows)

        print(f"\n{'=' * 60}")
        print(f"Combined CSV: {combined_name}")
        n_managers = len(MANAGERS_13F) + len(MANAGERS_NPORT)
        print(f"   {n_managers} managers, {len(combined_rows)} rows")
        print("=" * 60)
    else:
        print("\nNo data collected.")


if __name__ == "__main__":
    main()
